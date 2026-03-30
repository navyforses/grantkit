import openpyxl
import json
import re

all_resources = []
seen_orgs = set()

def clean(val):
    """Clean a cell value"""
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() in ["none", "nan", ""]:
        return ""
    return s

def extract_website(text):
    """Extract website URL from text"""
    if not text:
        return ""
    urls = re.findall(r'(?:https?://)?(?:www\.)?([a-zA-Z0-9-]+\.[a-zA-Z]{2,}(?:/[^\s,\n]*)?)', text)
    for u in urls:
        if '@' not in u and u not in ['gmail.com', 'outlook.com', 'yahoo.com']:
            return u
    return ""

def extract_email(text):
    """Extract email from text"""
    if not text:
        return ""
    emails = re.findall(r'[\w.+-]+@[\w-]+\.[\w.]+', text)
    return emails[0] if emails else ""

def extract_phone(text):
    """Extract phone from text"""
    if not text:
        return ""
    phones = re.findall(r'[\(]?\d{3}[\)]?[-.\s]?\d{3}[-.\s]?\d{4}', text)
    return phones[0] if phones else ""

def add_resource(org, category, description, website, email, phone, location, eligibility, notes, source_file):
    """Add a resource if not duplicate"""
    org = clean(org)
    if not org or len(org) < 3:
        return
    
    # Skip personal/Aleksandra-specific entries
    skip_words = ['ალექსანდრა', 'aleksandra', 'facebook', 'gofundme']
    if any(w in org.lower() for w in skip_words):
        return
    
    key = org.lower().replace(" ", "").replace("-", "")[:40]
    if key in seen_orgs:
        return
    seen_orgs.add(key)
    
    all_resources.append({
        "organization": org,
        "category": clean(category),
        "description": clean(description),
        "website": clean(website),
        "email": clean(email),
        "phone": clean(phone),
        "location": clean(location),
        "eligibility": clean(eligibility),
        "notes": clean(notes),
        "source": source_file
    })

# ============================================================
# FILE 1 & 2: Aleksandra_FULL_Funding_Guide
# ============================================================
for fname, label in [
    ("/home/ubuntu/upload/Aleksandra_FULL_Funding_Guide(3).xlsx", "guide_v1"),
    ("/home/ubuntu/upload/Aleksandra_FULL_Funding_Guide_UPDATED.xlsx", "guide_v2")
]:
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    
    # Duke EAP — დაფინანსება sheet
    ws = wb['Duke EAP — დაფინანსება']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if row and row[0] and isinstance(row[0], (int, float)):
            org = clean(row[1])
            cat = clean(row[2])
            desc = clean(row[3])
            web = clean(row[4])
            email_val = clean(row[5])
            phone_val = clean(row[6])
            b2 = clean(row[7]) if len(row) > 7 else ""
            amount = clean(row[8]) if len(row) > 8 else ""
            action = clean(row[9]) if len(row) > 9 else ""
            notes_val = clean(row[10]) if len(row) > 10 else ""
            
            email_final = extract_email(email_val) or extract_email(desc)
            phone_final = extract_phone(phone_val) or extract_phone(desc)
            web_final = extract_website(web)
            
            add_resource(org, cat, desc, web_final, email_final, phone_final, "", 
                        f"B-2 Visa: {b2}" if b2 else "", 
                        f"Amount: {amount}. {notes_val}" if amount else notes_val, label)
    
    # NAPA სია (304) sheet
    ws2 = wb['NAPA სია (304)']
    rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
    for row in rows2:
        if row and row[0] and row[1]:
            org = clean(row[1])
            prog = clean(row[2])
            web = clean(row[3])
            email_val = clean(row[4])
            phone_val = clean(row[5])
            cat = clean(row[6])
            relevance = clean(row[7]) if len(row) > 7 else ""
            notes_val = clean(row[8]) if len(row) > 8 else ""
            
            desc = f"{prog}" if prog else notes_val
            email_final = extract_email(email_val) or extract_email(phone_val)
            phone_final = extract_phone(phone_val) or extract_phone(email_val)
            web_final = extract_website(web)
            
            add_resource(org, cat, desc, web_final, email_final, phone_final, "", "", notes_val, f"napa_{label}")
    
    # რუკა & ანალიზი (only in updated version)
    if 'რუკა & ანალიზი' in wb.sheetnames:
        ws3 = wb['რუკა & ანალიზი']
        rows3 = list(ws3.iter_rows(min_row=5, values_only=True))
        for row in rows3:
            if row and row[1]:
                org = clean(row[1])
                rtype = clean(row[2])
                desc = clean(row[3])
                amount = clean(row[4])
                notes_val = clean(row[7]) if len(row) > 7 else ""
                add_resource(org, rtype, desc, "", "", "", "", "", 
                           f"Amount: {amount}. {notes_val}" if amount else notes_val, f"map_{label}")
    
    wb.close()

# ============================================================
# FILE 3: dukeსდაფინანსება.xlsx
# ============================================================
wb3 = openpyxl.load_workbook("/home/ubuntu/upload/dukeსდაფინანსება.xlsx", read_only=True, data_only=True)

for tier_name in ['ტიერ 1 - მაღალი', 'ტიერ 2 - საშუალო', 'ტიერ 3 - დაბალი']:
    ws = wb3[tier_name]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    for row in rows:
        if row and row[0] and row[1]:
            org = clean(row[1])
            desc = clean(row[2])
            amount = clean(row[3])
            contact = clean(row[4])
            b2 = clean(row[5])
            action = clean(row[6])
            
            email_final = extract_email(contact)
            phone_final = extract_phone(contact)
            web_final = extract_website(contact)
            
            add_resource(org, tier_name, desc, web_final, email_final, phone_final, "", 
                        f"B-2 Visa: {b2}" if b2 else "",
                        f"Amount: {amount}. {action}" if amount else action, "duke_funding")

# EAP გამოცდილება sheet
if '📋 EAP გამოცდილება' in wb3.sheetnames:
    ws_exp = wb3['📋 EAP გამოცდილება']
    rows_exp = list(ws_exp.iter_rows(min_row=2, values_only=True))
    for row in rows_exp:
        if row and row[0] and row[1]:
            org = clean(row[1]) if len(row) > 1 else ""
            if org:
                desc_parts = [clean(c) for c in row[2:] if clean(c)]
                add_resource(org, "EAP Experience", " | ".join(desc_parts[:3]), "", "", "", "", "", "", "duke_eap_exp")

wb3.close()

# ============================================================
# FILE 5: საცხოვრებელი.xlsx
# ============================================================
wb5 = openpyxl.load_workbook("/home/ubuntu/upload/საცხოვრებელი.xlsx", read_only=True, data_only=True)

# რესურსები ★ sheet (header at row 3: #, პრიორ., ორგანიზაცია, კატეგორია, ღირებულება, მისამართი, ტელეფონი, ელ-ფოსტა/ვებ, უფლებამოს., მიმართვა, B-2, შენიშვნები)
ws = wb5['რესურსები ★']
rows = list(ws.iter_rows(min_row=5, values_only=True))
for row in rows:
    if row and len(row) > 2 and row[2]:
        org = clean(row[2])
        cat = clean(row[3]) if len(row) > 3 else ""
        cost = clean(row[4]) if len(row) > 4 else ""
        addr = clean(row[5]) if len(row) > 5 else ""
        phone_val = clean(row[6]) if len(row) > 6 else ""
        email_web = clean(row[7]) if len(row) > 7 else ""
        elig = clean(row[8]) if len(row) > 8 else ""
        process = clean(row[9]) if len(row) > 9 else ""
        b2 = clean(row[10]) if len(row) > 10 else ""
        notes_val = clean(row[11]) if len(row) > 11 else ""
        
        email_final = extract_email(email_web)
        phone_final = extract_phone(phone_val) or extract_phone(email_web)
        web_final = extract_website(email_web)
        
        add_resource(org, cat, f"{cost}. {process}" if process else cost, web_final, email_final, phone_final, addr, 
                    f"{elig}. B-2: {b2}" if b2 else elig, notes_val, "housing_star")

# საცხოვრებელი რესურსები sheet (header: კატეგორია, ორგანიზაცია, მისამართი, ღირებულება, უფლებამოსილება, საკონტაქტო, მიმართვა, B-2, პრიორიტეტი)
ws2 = wb5['საცხოვრებელი რესურსები']
rows2 = list(ws2.iter_rows(min_row=5, values_only=True))
for row in rows2:
    if row and len(row) > 1 and row[1]:
        cat = clean(row[0])
        org = clean(row[1])
        addr = clean(row[2]) if len(row) > 2 else ""
        cost = clean(row[3]) if len(row) > 3 else ""
        elig = clean(row[4]) if len(row) > 4 else ""
        contact = clean(row[5]) if len(row) > 5 else ""
        process = clean(row[6]) if len(row) > 6 else ""
        b2 = clean(row[7]) if len(row) > 7 else ""
        notes_val = clean(row[8]) if len(row) > 8 else ""
        
        email_final = extract_email(contact)
        phone_final = extract_phone(contact)
        web_final = extract_website(contact)
        
        add_resource(org, cat, f"{cost}. {process}" if process else cost, web_final, email_final, phone_final, addr, 
                    f"{elig}. B-2: {b2}" if b2 else elig, notes_val, "housing_resources")

wb5.close()

# ============================================================
# CATEGORIZE AND CLEAN
# ============================================================
category_map = {
    "MEDICAL": "Medical & Treatment",
    "MEDS": "Medical & Treatment",
    "THERAPY": "Medical & Treatment",
    "სამედიცინო": "Medical & Treatment",
    "ტიერ 1 - მაღალი": "Medical & Treatment",
    "ტიერ 2 - საშუალო": "Medical & Treatment",
    "ტიერ 3 - დაბალი": "Medical & Treatment",
    "FINANCIAL": "Financial Assistance",
    "ფინანსური": "Financial Assistance",
    "SCHOLARSHIP": "Scholarships & Education",
    "AAC/COMM": "Assistive Technology & Communication",
    "EQUIPMENT": "Assistive Technology & Communication",
    "ADAPTIVE": "Assistive Technology & Communication",
    "HOUSING": "Housing & Accommodation",
    "საცხოვრებელი": "Housing & Accommodation",
    "საავადმყოფოსთან აფილირებული საცხოვრებელი": "Housing & Accommodation",
    "პაციენტის საცხოვრებელი": "Housing & Accommodation",
    "სამედიცინო ტარიფიანი სასტუმრო": "Housing & Accommodation",
    "არაკომერციული საცხოვრებელი": "Housing & Accommodation",
    "რელიგიაზე დაფუძნებული მხარდაჭერა": "Social Services & Support",
    "მასპინძელი ოჯახის ქსელი": "Housing & Accommodation",
    "TRAVEL": "Travel & Transportation",
    "ფრენა": "Travel & Transportation",
    "LEGAL": "Legal & Immigration",
    "OTHER": "Other Services",
    "Duke სერვისი": "Duke University Services",
    "Duke შიდა": "Duke University Services",
    "EAP Experience": "Duke University Services",
    "Crowdfunding": "Financial Assistance",
    "UK ფონდი": "International Foundations",
    "ფონდი": "International Foundations",
}

for r in all_resources:
    cat = r["category"]
    mapped = None
    for key, val in category_map.items():
        if key.lower() in cat.lower():
            mapped = val
            break
    if not mapped:
        if any(w in cat.lower() for w in ["medical", "health", "სამედიცინო"]):
            mapped = "Medical & Treatment"
        elif any(w in cat.lower() for w in ["housing", "საცხოვრებელი", "home"]):
            mapped = "Housing & Accommodation"
        elif any(w in cat.lower() for w in ["financial", "fund", "ფინანს"]):
            mapped = "Financial Assistance"
        else:
            mapped = "Other Services"
    r["category"] = mapped

# Add IDs
for i, r in enumerate(all_resources):
    r["id"] = f"res_{i+1:03d}"

print(f"Total unique resources extracted: {len(all_resources)}")
print(f"\nCategories:")
cats = {}
for r in all_resources:
    cats[r["category"]] = cats.get(r["category"], 0) + 1
for c, count in sorted(cats.items(), key=lambda x: -x[1]):
    print(f"  {c}: {count}")

# Save
with open("/home/ubuntu/grantkit/extracted_resources.json", "w", encoding="utf-8") as f:
    json.dump(all_resources, f, ensure_ascii=False, indent=2)

print(f"\nSaved to extracted_resources.json")
print(f"\nSample (first 3):")
for r in all_resources[:3]:
    print(f"  {r['id']}: {r['organization']} [{r['category']}]")
    print(f"    Desc: {r['description'][:80]}...")
    print(f"    Web: {r['website']} | Email: {r['email']} | Phone: {r['phone']}")
