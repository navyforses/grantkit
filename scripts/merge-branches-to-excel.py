#!/usr/bin/env python3
"""
Phase 3b-4: Merge Places API branches into existing Excel workbook.

Reads:
  data/organizations-2026-04-20.xlsx  (existing 538-org workbook)
  data/orgs-phase2.json               (consolidated DB data, for HQ fallback)
  data/orgs-phase3b.json              (Places API results)

Writes:
  data/organizations-2026-04-20.xlsx  (updated in-place — backup to .bak first)

Behaviour:
  - Organizations sheet: update "Branches Count" column (N=branches + 1 HQ).
                        Fill blank Latitude/Longitude from Places API HQ match.
  - Branches sheet:      rebuild. For each org:
                          row 1: HQ (from DB, fallback to Places API HQ if DB blank)
                          row 2..N: Places API branches
"""
import json
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = Path(__file__).resolve().parent.parent / "data"
XLSX = DATA_DIR / "organizations-2026-04-20.xlsx"
BAK = DATA_DIR / "organizations-2026-04-20.bak.xlsx"
PHASE2 = DATA_DIR / "orgs-phase2.json"
PHASE3B = DATA_DIR / "orgs-phase3b.json"

# Sandbox fallback
if not XLSX.exists():
    alt = Path("/sessions/modest-magical-goldberg/mnt/outputs/data")
    if (alt / "organizations-2026-04-20.xlsx").exists():
        DATA_DIR = alt
        XLSX = alt / "organizations-2026-04-20.xlsx"
        BAK = alt / "organizations-2026-04-20.bak.xlsx"
        PHASE2 = alt / "orgs-phase2.json"
        PHASE3B = alt / "orgs-phase3b.json"
    else:
        # Try repo path
        repo = Path("/sessions/modest-magical-goldberg/mnt/grantkit-main/data")
        XLSX = repo / "organizations-2026-04-20.xlsx"
        BAK = repo / "organizations-2026-04-20.bak.xlsx"
        PHASE2 = repo / "orgs-phase2.json"
        PHASE3B = repo / "orgs-phase3b.json"
        if not PHASE2.exists():
            PHASE2 = Path("/sessions/modest-magical-goldberg/mnt/outputs/data/orgs-phase2.json")
        if not PHASE3B.exists():
            PHASE3B = Path("/sessions/modest-magical-goldberg/mnt/outputs/data/orgs-phase3b.json")
        DATA_DIR = repo

print(f"XLSX:    {XLSX}")
print(f"PHASE2:  {PHASE2}")
print(f"PHASE3B: {PHASE3B}")

# Styling reused from Phase 3 build
FONT = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HQ_FILL = PatternFill("solid", fgColor="E7F3FF")  # light blue for HQ rows
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def main():
    with open(PHASE2, "r", encoding="utf-8") as f:
        phase2 = json.load(f)
    with open(PHASE3B, "r", encoding="utf-8") as f:
        phase3b = json.load(f)

    # Index phase3b by org_id
    p3b_by_id = {r["org_id"]: r for r in phase3b}

    # Backup
    if XLSX.exists():
        shutil.copy2(XLSX, BAK)
        print(f"Backup saved: {BAK}")

    wb = load_workbook(XLSX)

    # ----- Organizations sheet update -----
    orgs_sheet = wb["Organizations"]
    header = [c.value for c in orgs_sheet[1]]
    print(f"Organizations headers: {header}")

    # Locate columns we need to update
    col_map = {name: idx + 1 for idx, name in enumerate(header)}
    col_lat = col_map.get("Latitude")
    col_lng = col_map.get("Longitude")
    col_branches = col_map.get("Branches Count")

    for i, org in enumerate(phase2):
        row = i + 2  # data starts at row 2
        org_id = f"ORG-{i+1:04d}"
        p3b = p3b_by_id.get(org_id)

        # Fill missing lat/lng from Places API HQ
        if p3b and p3b.get("hq"):
            hq = p3b["hq"]
            cur_lat = orgs_sheet.cell(row=row, column=col_lat).value
            cur_lng = orgs_sheet.cell(row=row, column=col_lng).value
            if (cur_lat is None or cur_lat == "") and hq.get("latitude"):
                orgs_sheet.cell(row=row, column=col_lat).value = hq["latitude"]
            if (cur_lng is None or cur_lng == "") and hq.get("longitude"):
                orgs_sheet.cell(row=row, column=col_lng).value = hq["longitude"]

        # Update Branches Count: HQ (1) + Places API branches
        branches = (p3b or {}).get("branches") or []
        orgs_sheet.cell(row=row, column=col_branches).value = 1 + len(branches)

    # ----- Branches sheet rebuild -----
    # Delete existing Branches sheet, recreate
    if "Branches" in wb.sheetnames:
        del wb["Branches"]
    br = wb.create_sheet("Branches")

    br_headers = [
        "org_id", "branch_id", "Organization", "Branch Type",
        "Country", "State", "City", "Address",
        "Phone", "Email", "Latitude", "Longitude",
        "Source", "Notes",
    ]
    for c, h in enumerate(br_headers, 1):
        cell = br.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    br.column_dimensions["A"].width = 11
    br.column_dimensions["B"].width = 13
    br.column_dimensions["C"].width = 40
    br.column_dimensions["D"].width = 12
    br.column_dimensions["E"].width = 8
    br.column_dimensions["F"].width = 6
    br.column_dimensions["G"].width = 18
    br.column_dimensions["H"].width = 45
    br.column_dimensions["I"].width = 18
    br.column_dimensions["J"].width = 25
    br.column_dimensions["K"].width = 12
    br.column_dimensions["L"].width = 12
    br.column_dimensions["M"].width = 14
    br.column_dimensions["N"].width = 30

    br.freeze_panes = "A2"
    br.row_dimensions[1].height = 28

    row = 2
    for i, org in enumerate(phase2):
        org_id = f"ORG-{i+1:04d}"
        p3b = p3b_by_id.get(org_id)

        # -- HQ row --
        # Prefer DB address if present, else Places API HQ
        hq_from_db = bool(org.get("address"))
        if hq_from_db:
            hq_data = {
                "address": org.get("address", ""),
                "state": org.get("state", ""),
                "city": org.get("city", ""),
                "phone": org.get("phone", ""),
                "email": org.get("email", ""),
                "latitude": org.get("latitude", ""),
                "longitude": org.get("longitude", ""),
                "source": "Database",
            }
        elif p3b and p3b.get("hq"):
            hq = p3b["hq"]
            hq_data = {
                "address": hq.get("formatted_address", ""),
                "state": "",
                "city": "",
                "phone": hq.get("phone", "") or "",
                "email": "",
                "latitude": hq.get("latitude", ""),
                "longitude": hq.get("longitude", ""),
                "source": "Google Places (HQ)",
            }
        else:
            hq_data = {
                "address": "", "state": "", "city": "",
                "phone": "", "email": "",
                "latitude": "", "longitude": "",
                "source": "Not found",
            }

        row_vals = [
            org_id,
            f"{org_id}-B01",
            org.get("organization", ""),
            "HQ",
            org.get("country", ""),
            hq_data["state"],
            hq_data["city"],
            hq_data["address"],
            hq_data["phone"],
            hq_data["email"],
            hq_data["latitude"],
            hq_data["longitude"],
            hq_data["source"],
            "Headquarters" if hq_from_db else ("From Google Places" if p3b and p3b.get("hq") else "No HQ address in DB or Places API"),
        ]
        for c, v in enumerate(row_vals, 1):
            cell = br.cell(row=row, column=c, value=v)
            cell.font = FONT
            cell.fill = HQ_FILL
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

        # -- Branch rows from Places API --
        branches = (p3b or {}).get("branches") or []
        for b_idx, b in enumerate(branches, 2):
            addr = b.get("formatted_address", "") or ""
            # Parse city from addressComponents if present, else extract from address
            city, state = "", ""
            # Simple: address components not carried; use empty — Google Places formatted_address contains it
            row_vals = [
                org_id,
                f"{org_id}-B{b_idx:02d}",
                b.get("name", "") or org.get("organization", ""),
                "Branch",
                b.get("country", "") or org.get("country", ""),
                state,
                city,
                addr,
                b.get("phone", "") or "",
                "",
                b.get("latitude", ""),
                b.get("longitude", ""),
                "Google Places",
                f"Type: {b.get('primary_type', '') or ''}",
            ]
            for c, v in enumerate(row_vals, 1):
                cell = br.cell(row=row, column=c, value=v)
                cell.font = FONT
                cell.border = BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            row += 1

    # ----- Update Summary sheet -----
    if "Summary" in wb.sheetnames:
        s = wb["Summary"]
        # Look for existing branch count cell or append
        # Update a generic info line at the bottom
        last_row = s.max_row
        s.cell(row=last_row + 2, column=1, value="Branches enriched via Google Places API (New) — Phase 3b").font = Font(name="Arial", size=10, bold=True)
        s.cell(row=last_row + 3, column=1, value="Total branch rows:")
        s.cell(row=last_row + 3, column=2, value=row - 2 - len(phase2)).font = Font(name="Arial", size=10, bold=True)
        s.cell(row=last_row + 4, column=1, value="Orgs with 1+ real branches:")
        orgs_with_br = sum(1 for r in phase3b if r.get("branches"))
        s.cell(row=last_row + 4, column=2, value=orgs_with_br).font = Font(name="Arial", size=10, bold=True)
        s.cell(row=last_row + 5, column=1, value="HQ match rate:")
        ok = sum(1 for r in phase3b if r["search_status"] == "ok")
        s.cell(row=last_row + 5, column=2, value=f"{ok}/{len(phase3b)} ({ok/len(phase3b)*100:.1f}%)").font = Font(name="Arial", size=10, bold=True)

    wb.save(XLSX)
    print(f"\nSaved: {XLSX}")
    print(f"Branch rows written: {row - 2}")
    print(f"  HQ rows:     {len(phase2)}")
    print(f"  Branch rows: {row - 2 - len(phase2)}")


if __name__ == "__main__":
    main()
